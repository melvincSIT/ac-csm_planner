#!/usr/bin/env python3
"""Independent verification of data.json against Excel v2, SIT, and scheduling rules."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Run: python3 -m venv .venv && .venv/bin/pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent
RAW = json.loads((ROOT / "data.json").read_text())
REGISTRY = RAW if "programs" in RAW else {"programs": {"applied-computing": RAW}}
PROGRAMS = REGISTRY["programs"]
WB = openpyxl.load_workbook(ROOT / "Candidature Plan_30 Jun 26 v2.xlsx", data_only=True)

EXCEL_NAME_TO_ID = {
    "Digital Logic & Computing": "dlc",
    "Ethical Computing & Data Analytics": "ethical",
    "Information Security Management & Audit": "isma",
    "Modern Software Engineering Practices": "mse",
    "Cloud Architecting & Security": "cloud",
    "Machine Learning": "ml",
    "Computer Networking & Network Security": "netsec",
    "Cybersecurity": "cyber",
    "Advanced Generative Text Artificial Intelligence": "genai",
    "Deep Learning Forecasting with Time Series Analysis": "dlts",
}

errors = []
warnings = []


def feature_on(prog, name):
    feats = prog.get("features") or {}
    if name in ("rpl", "reattempt", "remodule"):
        return feats.get(name) is True
    return feats.get(name, True) is not False


def planner_credit_total(prog):
    mc = prog["mcCount"] * prog["mcCredits"]
    wfe = prog.get("wfeCcCredits", 0) if feature_on(prog, "wfeCc") else 0
    riwe = prog.get("riweCredits", 0) if feature_on(prog, "riwe") else 0
    cap = prog.get("capstoneCredits", 0) if feature_on(prog, "capstone") else 0
    return mc + wfe + riwe + cap


def validate_program(pid, prog):
    title = prog.get("degreeTitle") or prog.get("title") or pid
    planner_total = planner_credit_total(prog)
    if prog.get("totalCredits") != planner_total:
        errors.append(
            f"{title}: totalCredits {prog.get('totalCredits')} != planner sum {planner_total}"
        )

    full = prog.get("fullDegreeCredits")
    if full is not None and full < prog.get("totalCredits", 0):
        errors.append(f"{title}: fullDegreeCredits ({full}) < totalCredits")

    if prog.get("maxCoreTrimesters", 0) + prog.get("maxExtensionTrimesters", 0) != prog.get(
        "maxTotalTrimesters", 0
    ):
        errors.append(
            f"{title}: maxCoreTrimesters + maxExtensionTrimesters != maxTotalTrimesters"
        )

    courses = prog.get("courses") or []
    course_ids = {c["id"] for c in courses}
    mc_courses = [c for c in courses if c.get("category") in ("foundation", "stackable")]
    foundations = [c for c in mc_courses if c.get("category") == "foundation"]
    stackables = [c for c in mc_courses if c.get("category") == "stackable"]

    if len(foundations) < prog.get("foundationCount", 0):
        errors.append(f"{title}: only {len(foundations)} foundation courses in catalogue")
    if len(stackables) < prog.get("stackableCount", 0):
        errors.append(f"{title}: only {len(stackables)} stackable courses in catalogue")

    for month, ids in (prog.get("offeringsByMonth") or {}).items():
        for cid in ids:
            if cid not in course_ids:
                errors.append(f"{title}: offeringsByMonth[{month}] references unknown id {cid!r}")

    cycle = prog.get("monthCycle") or []
    for intake_key, info in (prog.get("intakes") or {}).items():
        if info.get("startMonth") not in cycle:
            errors.append(f"{title}: intake {intake_key} startMonth not in monthCycle")
        for t in info.get("wfeRiweTrimesters") or []:
            if feature_on(prog, "riwe"):
                year = (t + 2) // 3
                riwe_year = (prog.get("constraints") or {}).get("wfeRiweYear")
                if riwe_year and year != riwe_year:
                    errors.append(
                        f"{title}: RIWE trimester {t} for {intake_key} is year {year}, expected {riwe_year}"
                    )
        for t in info.get("capstoneTrimesters") or []:
            if feature_on(prog, "capstone"):
                year = (t + 2) // 3
                cap_year = (prog.get("constraints") or {}).get("capstoneMinYear")
                if cap_year and year < cap_year:
                    errors.append(
                        f"{title}: Capstone trimester {t} for {intake_key} is year {year}, min {cap_year}"
                    )

    if pid == "applied-computing":
        validate_applied_computing_excel(prog)


def validate_applied_computing_excel(DATA):
    mc_total = DATA["mcCount"] * DATA["mcCredits"]
    expected = mc_total + DATA["wfeCcCredits"] + DATA["capstoneCredits"]
    if DATA.get("riweCredits"):
        expected += DATA["riweCredits"]
    if expected != DATA["totalCredits"]:
        errors.append(f"Applied Computing: credit sum should be {expected}, got {DATA['totalCredits']}")

    ws = WB["I started in Sep"]
    excel_by_month = {"Sep": [], "Jan": [], "May": []}
    for col, month in [(3, "Sep"), (4, "Jan"), (5, "May")]:
        for r in range(27, 32):
            v = ws.cell(r, col).value
            if not v:
                continue
            s = str(v).strip()
            for ename, cid in EXCEL_NAME_TO_ID.items():
                if s.startswith(ename[:28]) or ename.startswith(s[:28]):
                    if cid not in excel_by_month[month]:
                        excel_by_month[month].append(cid)
                    break

    for month, ids in DATA["offeringsByMonth"].items():
        if set(ids) != set(excel_by_month.get(month, [])):
            errors.append(
                f"Applied Computing offeringsByMonth[{month}]: json={ids} excel={excel_by_month.get(month)}"
            )

    cycle = DATA["monthCycle"]
    for intake_key, info in DATA["intakes"].items():
        start = cycle.index(info["startMonth"])
        for t in range(1, 10):
            month = cycle[(start + t - 1) % 3]
            start_col = {"sep": 3, "jan": 4, "may": 5}[intake_key]
            col = start_col + t - 1
            excel_ids = []
            for r in range(27, 32):
                v = ws.cell(r, col).value if intake_key == "sep" else WB[
                    f"I started in {intake_key.capitalize()}"
                ].cell(r, col).value
                if not v:
                    continue
                s = str(v).strip()
                for ename, cid in EXCEL_NAME_TO_ID.items():
                    if s.startswith(ename[:28]) or ename.startswith(s[:28]):
                        if cid not in excel_ids:
                            excel_ids.append(cid)
                        break
            json_ids = [c for c in DATA["offeringsByMonth"][month]]
            if set(excel_ids) != set(json_ids):
                errors.append(
                    f"Applied Computing tri {t} {intake_key} ({month}): excel={excel_ids} json={json_ids}"
                )

    INTAKE_COL = {"sep": 3, "jan": 4, "may": 5}
    for key, start_col in INTAKE_COL.items():
        ws_intake = WB[f"I started in {key.capitalize()}"]
        for i in range(9):
            excel_val = ws_intake.cell(18, start_col + i).value
            json_val = DATA["intakes"][key]["idealFirstAttempt"][i]
            if excel_val != json_val:
                errors.append(
                    f"Ideal path {key} tri{i+1}: excel={excel_val!r} json={json_val!r}"
                )


print("=== CSM Planner Verification ===\n")
print(f"Programmes in registry: {len(PROGRAMS)}\n")

for pid, prog in sorted(PROGRAMS.items(), key=lambda x: x[1].get("order", 99)):
    validate_program(pid, prog)
    pt = planner_credit_total(prog)
    print(
        f"✓ checked {prog.get('degreeTitle', pid)}: "
        f"{prog['mcCount']}×{prog['mcCredits']} + components = {pt} cr "
        f"(target {prog['totalCredits']})"
    )

print()

if errors:
    print(f"ERRORS ({len(errors)}):")
    for e in errors:
        print("  ✗", e)
else:
    print("✓ All programme checks passed.\n")

warnings.append("Excel cross-check applies to Applied Computing only.")
warnings.append("Engineering programmes: planner tracks CSM pathway credits (138); full BEng is 240 cr.")
warnings.append("Excel: 'Management & Audit' vs SIT: 'Risk Management and Audit'.")
print("NOTES:")
for w in warnings:
    print("  ·", w)

sys.exit(1 if errors else 0)
